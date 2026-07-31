import openpyxl
from openpyxl import load_workbook
import glob
import os
import sys

FOLDER = "C:/Users/yonat/OneDrive/Desktop/memory/PC-06/docs/Absensi T.P 2026-2027"
MASTER = os.path.join(FOLDER, "ABSENSI Juli.xlsx")
SKIP = {"ABSENSI Juli.xlsx", "DAFTAR MURID T.P.2026-2027 lengkap.xlsx"}


def load_marks(path):
    wb = load_workbook(path)
    marks = {}
    for sn in wb.sheetnames:
        ws = wb[sn]
        datecol = {}
        for c in range(4, 50):
            t = ws.cell(6, c).value
            if t and str(t).isdigit():
                datecol[int(t)] = c
        for row in range(7, 100):
            n = ws.cell(row, 3).value
            if not n or len(str(n).strip()) <= 1:
                continue
            name = str(n).strip()
            for d, c in datecol.items():
                v = ws.cell(row, c).value
                if v not in (None, ""):
                    marks[(sn, name.lower(), d)] = str(v)
    return marks


def main(target=MASTER):
    files = [f for f in glob.glob(os.path.join(FOLDER, "*.xlsx"))
             if os.path.basename(f) not in SKIP]
    files.append(MASTER)

    all_marks = {}
    for f in files:
        m = load_marks(f)
        print(f"[LOAD] {os.path.basename(f)} -> {len(m)} tanda")
        all_marks.update(m)

    print(f"[TOTAL] {len(all_marks)} tanda unik akan digabung")

    wb = load_workbook(target)
    added = 0
    for sn in wb.sheetnames:
        ws = wb[sn]
        datecol = {}
        for c in range(4, 50):
            t = ws.cell(6, c).value
            if t and str(t).isdigit():
                datecol[int(t)] = c
        name2row = {}
        for row in range(7, 100):
            n = ws.cell(row, 3).value
            if n and len(str(n).strip()) > 1:
                name2row[str(n).strip().lower()] = row
        for (sname, nlow, d), v in all_marks.items():
            if sname != sn or nlow not in name2row:
                continue
            row = name2row[nlow]
            col = datecol.get(d)
            if col is None:
                continue
            cur = ws.cell(row, col).value
            if cur in (None, ""):
                ws.cell(row, col).value = v
                added += 1
            elif str(cur).strip() != str(v).strip():
                print(f"  !CONFLICT {sn} | {nlow} | tgl {d}: master={cur}, lain={v}")

    wb.save(target)
    print(f"[OK] Selesai. {added} tanda baru ditambahkan ke {os.path.basename(target)}")


if __name__ == "__main__":
    t = sys.argv[1] if len(sys.argv) > 1 else MASTER
    main(t)
