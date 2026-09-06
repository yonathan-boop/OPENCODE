import openpyxl
import argparse
import os
import sys

DEFAULT_FOLDER = os.path.join("PC-06", "docs", "Absensi T.P 2026-2027")


BULAN = ["Januari", "Februari", "Maret", "April", "Mei", "Juni",
         "Juli", "Agustus", "September", "Oktober", "November", "Desember"]


def parse_name(name):
    toks = name.replace(".xlsx", "").split()
    if len(toks) < 3 or toks[0].lower() != "absensi":
        return None
    year = None
    month = None
    day = 0
    hm = "000000"
    for t in toks:
        if t.isdigit() and len(t) == 4:
            year = int(t)
        if t in BULAN:
            month = BULAN.index(t)
    if tok2 := toks[1]:
        if tok2.isdigit():
            day = int(tok2)
    for t in toks:
        if t.count("_") == 2 and t.replace("_", "").isdigit():
            hm = t.replace("_", "")
    if year is None or month is None:
        return None
    return (year, month, day, hm)


def find_latest(folder):
    parsed = []
    for f in os.listdir(folder):
        if not f.lower().endswith(".xlsx"):
            continue
        key = parse_name(f)
        if key:
            parsed.append((key, f))
    if not parsed:
        sys.exit(f"[ERROR] Tidak ada file absensi valid di {folder}")
    parsed.sort(key=lambda x: x[0])
    return parsed[-1][1]


def build_tanggal_map(ws):
    m = {}
    for c in range(4, 50):
        v = ws.cell(6, c).value
        if v is not None and str(v).isdigit():
            m[int(v)] = c
    return m


def rekap(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    tahun = parse_name(os.path.basename(path))
    tahun = tahun[0] if tahun else ""
    print(f"REKAP ABSENSI — {os.path.basename(path)}")
    print("=" * 78)
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        bulan = ws.cell(5, 4).value or bulan
        tgl_map = build_tanggal_map(ws)
        per_day = {t: {"S": 0, "I": 0, "A": 0} for t in tgl_map}
        siswa_absen = {}
        total_siswa = 0
        for r in range(7, ws.max_row + 1):
            nama = ws.cell(r, 3).value
            if not nama or len(str(nama).strip()) <= 1:
                continue
            total_siswa += 1
            daftar = []
            for t, c in tgl_map.items():
                v = ws.cell(r, c).value
                if v in ("S", "I", "A"):
                    per_day[t][v] += 1
                    daftar.append(f"{t}-{v}")
            if daftar:
                s = sum(1 for d in daftar if d.endswith("-S"))
                i = sum(1 for d in daftar if d.endswith("-I"))
                a = sum(1 for d in daftar if d.endswith("-A"))
                siswa_absen[str(nama).strip()] = (s, i, a, daftar)
        print(f"\n[{sheet}] — {bulan} {tahun} | total murid: {total_siswa}")
        print(f"  Ringkasan harian:")
        for t in sorted(per_day):
            d = per_day[t]
            if any(d.values()):
                print(f"    {bulan} {t:>2}: S={d['S']} I={d['I']} A={d['A']}")
        if not siswa_absen:
            print("  (belum ada absensi)")
            continue
        print(f"  Murid absen (terbanyak ke bawah):")
        for nama, (s, i, a, daftar) in sorted(siswa_absen.items(), key=lambda x: -(x[1][0] + x[1][1] + x[1][2])):
            tot = s + i + a
            print(f"    {nama}: {tot}x -> " + ", ".join(daftar))


if __name__ == "__main__":
    ap = argparse.ArgumentParser(description="Rekap absensi bulanan dari file Excel versi terbaru")
    ap.add_argument("--folder", default=DEFAULT_FOLDER, help="Folder berisi file absensi (default: relatif ke repo)")
    ap.add_argument("--file", default=None, help="Nama file spesifik (default: file absensi terbaru di folder)")
    args = ap.parse_args()
    path = args.file if args.file else os.path.join(args.folder, find_latest(args.folder))
    if not os.path.exists(path):
        sys.exit(f"[ERROR] File tidak ditemukan: {path}")
    rekap(path)