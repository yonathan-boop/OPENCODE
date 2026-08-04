import openpyxl
from openpyxl import load_workbook
import sys

FILE_PATH = "C:/Users/yonat/OneDrive/Desktop/memory/PC-06/docs/Absensi T.P 2026-2027/ABSENSI Agustus.xlsx"

KELAS_MAP = {
    'tka': 'TKa', 'tk a': 'TKa',
    'tkb1': 'TKB1', 'tk b1': 'TKB1',
    'tkb2': 'TKB(2)', 'tk b2': 'TKB(2)', 'tkb(2)': 'TKB(2)',
    'pg': 'Absen PG', 'playgroup': 'Absen PG', 'absen pg': 'Absen PG'
}

KODE = {
    's': 'S', 'sakit': 'S',
    'i': 'I', 'izin': 'I',
    'a': 'A', 'alpha': 'A',
    '.': '.'
}

def build_tanggal_map(ws):
    tanggal_to_col = {}
    for c in range(4, 50):
        tgl = ws.cell(6, c).value
        if tgl and str(tgl).isdigit():
            tanggal_to_col[int(tgl)] = c
    return tanggal_to_col

def absen(nama, kelas, tanggal, alasan):
    wb = load_workbook(FILE_PATH)

    kelas = kelas.lower().strip()
    sheet_name = KELAS_MAP.get(kelas)

    if not sheet_name:
        print(f"[ERROR] Kelas tidak dikenal: {kelas}")
        return

    ws = wb[sheet_name]
    tanggal_map = build_tanggal_map(ws)

    if int(tanggal) not in tanggal_map:
        print(f"[ERROR] Tanggal {tanggal} tidak ada di sheet {sheet_name}")
        print(f"Tanggal tersedia: {sorted(tanggal_map.keys())}")
        return

    col_tanggal = tanggal_map[int(tanggal)]
    kode = KODE.get(alasan.lower(), alasan.upper())

    row_ditemukan = None
    nama_ditemukan = None
    for row in range(7, 100):
        nama_murid = ws.cell(row, 3).value
        if not nama_murid or len(str(nama_murid).strip()) <= 1:
            continue
        if nama.lower() in nama_murid.lower() or nama_murid.lower() in nama.lower():
            row_ditemukan = row
            nama_ditemukan = nama_murid
            break

    if not row_ditemukan:
        print(f"[ERROR] Murid '{nama}' tidak ditemukan di kelas {kelas}")
        return

    ws.cell(row_ditemukan, col_tanggal).value = kode
    wb.save(FILE_PATH)

    print(f"[OK] {nama_ditemukan} ({sheet_name}) - Tanggal {tanggal} = {kode}")

if __name__ == "__main__":
    if len(sys.argv) < 5:
        print("Cara pakai: py absensi.py <nama> <kelas> <tanggal> <alasan>")
        print("Contoh: py absensi.py Aldrich TKa 18 sakit")
        sys.exit(1)

    absen(sys.argv[1], sys.argv[2], sys.argv[3], sys.argv[4])